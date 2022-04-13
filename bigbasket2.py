#creating an excel sheet which contains the data to be searched
# from openpyxl import *
# from  openpyxl.utils import get_column_letter
# from openpyxl.styles import Font
# data=["kissan fruit jam","dabour honey",'ashirwad atta']
# wb=Workbook() #creating a new excel workbook
# ws=wb.active #creating a worksheet
# headings=["PRODUCTS"]
# ws.append(headings)
# for i in data:
#     ws.append([i])
# wb.save("products_to_search.xlsx")

# reading the excel values and making into a list
import openpyxl
from  openpyxl.utils import get_column_letter
from openpyxl.styles import Font
wb=openpyxl.load_workbook("products_to_search.xlsx")
ws=wb.active
rows=ws.max_row
cols=ws.max_column
# print(rows,cols)
pro_list=[]
for r in range(2,rows+1):
    for c in range(1,cols+1):
        # print(ws.cell(row=r,column=c).value)
        x=ws.cell(row=r,column=c).value
        pro_list.append(x)
print(pro_list)

#entering into the bigbasket website through web automation
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.select import Select
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.keys import Keys
import keyboard
import datetime
import time
from selenium.webdriver.common.action_chains import ActionChains
import re
import openpyxl
from  openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl import *
wb=Workbook()
ws=wb.active
headings=["Product_name","rating","Quantity","Rate in Rs"]
ws.append(headings)

path="C:\Program Files (x86)\chromedriver.exe"
driver=webdriver.Chrome(path)
driver.get("https://www.bigbasket.com/")
driver.maximize_window()
complete_data=""
for i in pro_list:
    search_bar=driver.find_element_by_xpath("//*[@id='input']").send_keys(i,Keys.ENTER)
    time.sleep(5)
    driver.execute_script("window.scrollTo(0, 300);")
    try:
        eles=driver.find_elements_by_class_name("items")
        all_data=""
        for e in eles:
            e=e.text
            all_data=all_data+e
            all_data.split("\n")
        complete_data=complete_data+all_data
        print('--------------------------------------------')
        print(complete_data)
    

    except:
        print("no items found")
# print(complete_data)
pro_data=re.findall("(?s)Kissan.*?ADD|(?s)Dabur.*?ADD|(?s)Aashirvaad.*?ADD|(?s)Kissan.*?NOTIFY ME|(?s)Aashirvaad.*?NOTIFY ME|(?s)Dabur.*?NOTIFY ME", complete_data)
# print(pro_data)
for i in pro_data:
    i=i.split("\n")
    print(i)
    # print("-------------",len(i))
    product_name=i[0]+"_"+i[1]
    rating=i[2]
    quantity=i[4].split("-")[0]
    if len(i)==8 and i[-1]!='NOTIFY ME':
        quantity=i[2]
        rating="NA"
    price=i[-4].split("Rs")[-1]
    if i[-1]=='NOTIFY ME':
        price=i[-2].split("Rs")[-1]
    if len(i)==7:
       quantity=i[1] 
       rating="NA" 
    list1=[product_name,rating,quantity,price]
    print(list1)
    ws.append(list1)
wb.save("bigbasket3.xlsx")
driver.quit()














    
