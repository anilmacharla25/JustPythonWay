from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.select import Select
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.keys import Keys
import keyboard
import datetime
import time
from selenium.webdriver.common.action_chains import ActionChains
import re
import requests
import io
from PIL import Image

from openpyxl import *
from  openpyxl.utils import get_column_letter
from openpyxl.styles import Font
wb=Workbook() #creating new excel workbok
ws=wb.active
ws.title="blackbaghr"


options = Options()

options.add_experimental_option('excludeSwitches', ['enable-logging'])

path="C:\Program Files (x86)\chromedriver.exe"
driver=webdriver.Chrome(path)
driver.get("https://dermnetnz.org/image-library/")
driver.maximize_window()

'''-------------get name of desease----------------'''
desease_list=[]
names=driver.find_elements_by_class_name("imageList__group__item__copy")
for name in names:
    # print(name.text)
    desease_list.append(name.text)
print(desease_list)
time.sleep(1)
print(len(desease_list))
'''------------------url assosiated with desease-------------'''
urls_list=[]
urls=driver.find_elements(By.TAG_NAME,"a")
# print(len(urls))
for url in urls:
    link_text=url.get_attribute("href")
    try:
        if '=Live' in link_text:
            print(link_text)
            # time.sleep(1)
            try:
                driver.implicitly_wait(30)
                urls_list.append(link_text)
            except:
                pass
    except:
        pass
print(urls_list)
print(len(urls_list))

'''----------------for icon images----------------'''
image_urls_list=[]
image_urls=driver.find_elements(By.TAG_NAME,"img")
print(len(image_urls))
for url in image_urls:
    
    try:
        link_text=url.get_attribute("src")
        print(link_text)
        if '.png' in link_text:
            pass
        elif 'None' in link_text:
            pass
        else:
            print(link_text)
            driver.implicitly_wait(30)
            image_urls_list.append(link_text)
    except:
        pass
time.sleep(1)
print(image_urls_list)

# for url in image_urls_list:

#     driver.implicitly_wait(30)
#     print('ok')
#     downlod_path="C:\\Users\\prave\\Desktop\\DATA_SET\\imgs\\"
#     i=0
#     file_name=url[38:-4]+".jpg"
#     i=i+1
#     try:
#         image_content=requests.get(url).content
#         image_file=io.BytesIO(image_content)
#         image=Image.open(image_file)
#         file_path=downlod_path+file_name
#         print(file_path)

#         with open(file_path,"wb") as f:
#             image.save(f,"JPEG")
#             print('saved ')
           
        
#         print("sucess")
#     except Exception as e:
#         print("failed ",e)



headings=["Name of desease"]+["desease url"]+["img url"]
ws.append(headings)
for i in range(len(desease_list)):
    ws["A"+str(i+2)]=desease_list[i]
for i in range(len(urls_list)):
    ws["B"+str(i+2)]=urls_list[i]
for i in range(len(image_urls_list)):
    ws["C"+str(i+2)]=image_urls_list[i]

for col in range(1,5):
    ws[get_column_letter(col)+"1"].font=Font(bold=True,color="0099CCFF")
wb.save("newexel.xlsx")

    







