
#entering into site and scrapping the data
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.select import Select
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.keys import Keys
import keyboard
# import datetime
import time
from selenium.webdriver.common.action_chains import ActionChains
path="C:\Program Files (x86)\chromedriver.exe"
driver=webdriver.Chrome(path)
driver.get("https://www.mygov.in/covid-19/")
driver.maximize_window()
driver.execute_script("window.scrollTo(0, 600);")
time.sleep(2)
state_wise_dd=driver.find_elements_by_class_name('plus_icon')
state_wise_dd[1].click()
t_header=driver.find_element_by_xpath('//*[@id="ind_mp_tbl"]/thead/tr').text
t_header=t_header.split()
print(t_header)
header_list=[t_header[0],t_header[1]+t_header[2],t_header[3],t_header[4],t_header[5],t_header[6]+t_header[7],t_header[8]+t_header[9],t_header[10]+t_header[11]]
print(header_list)
# t_row=driver.find_element_by_xpath('//*[@id="ind_mp_tbl"]/tbody/tr[1]').text
# print(t_row.split("\n"))
rows=len(driver.find_elements_by_xpath("//*[@id='ind_mp_tbl']/tbody/tr"))
cols=len(driver.find_elements_by_xpath("//*[@id='ind_mp_tbl']/thead/tr/th"))
print(rows)
print(cols)
main_list=[]
for r in range(1,rows+1):
    data_list=[]
    for c in range(1,cols+1):
        value=driver.find_element_by_xpath(f'//*[@id="ind_mp_tbl"]/tbody/tr[{r}]/td[{c}]').text
        # print(value)
        if c>1:
            value=value.split()[0]
#             # print(value)
            if ',' in value:
                new_value=value.split(',')
                value="".join(new_value)
            data_list.append(value)
        else:
            # print(value)
            
            data_list.append(value)
    print()
    # print(data_list)
    main_list.append(data_list)
print(main_list)
driver.quit()
#creating an excel sheet and saving the data into excel sheet
from openpyxl import *
from  openpyxl.utils import get_column_letter
from openpyxl.styles import Font
wb=Workbook() #creating a new excel workbook
ws=wb.active #creating a worksheet
ws.append(header_list)
for i in main_list:
    ws.append(i)
wb.save("main_data.xlsx")
print("main list data has been added into main excel sheet sucsessfully")
print("--------------------------------------------------------------------------------------")
#creating other three tables in new excel sheet
# importing pandas module
import pandas as pd
import numpy as np

	
# making data frame
df = pd.read_excel('main_data.xlsx')

df.head(37)
most_likely = df.sort_values('TotalCases', ascending=False).head(5)
# print(most_likely)
table1=most_likely.drop(['Active', 'Discharged','Deaths','ActiveRatio','DeathRatio'], axis = 1)
print(table1)
#making one more dataframe
most_likely1=df.sort_values("Deaths", ascending=False).head(5)
table2=most_likely1.drop(['TotalCases', 'Discharged','ActiveRatio','DischargeRatio','DeathRatio'], axis = 1)
print(table2)
#summing the values in one column
from openpyxl import load_workbook
wb = load_workbook(filename="main_data.xlsx")
ws= wb.active
i=0
total_deaths=0
for cell in ws['E']:
    if i==0:
        pass
    else:
        # print(cell.value)
        total_deaths=total_deaths+int(cell.value)
    i+=1
    
print(total_deaths)
# lst = [total_deaths]
 
# Calling DataFrame constructor on list
# df3 = pd.DataFrame(lst)
# print(df3)
data = {'Total deaths in india':[total_deaths]}
df3 = pd.DataFrame(data)
print(df3)

# create excel writer object
writer = pd.ExcelWriter('sheet2.xlsx')
table1.to_excel(writer,sheet_name='Sheet',startrow=1 , startcol=0)
table2.to_excel(writer,sheet_name='Sheet',startrow=12 , startcol=0)
df3.to_excel(writer,sheet_name='Sheet',startrow=20 , startcol=0)
writer.save()
print(' three DataFrame is written successfully to Excel File.')